from urllib.request import urlopen, Request
from bs4 import BeautifulSoup
import pandas as pd
import requests
from openpyxl import load_workbook
import random

filename = 'Colleges.xlsx'

df = pd.read_excel(filename)

colleges = df['College'].tolist()

print(colleges)

acceptance_rates = []
for college in colleges:
    query = college + " acceptance rate"
    query = query.replace(' ', '+')
    URL = f"https://google.com/search?q={query}"
    USER_AGENT = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.14; rv:65.0) Gecko/20100101 Firefox/65.0"
    headers = {"user-agent" : USER_AGENT}
    resp = requests.get(URL, headers=headers)
    if resp.status_code == 200:
        soup = BeautifulSoup(resp.content, "html.parser")
        results = []
        for i in soup.find_all('div', {'class': 'Z0LcW'}):
            acceptance_rates.append(float(i.text[0:-1]))

print(acceptance_rates)

book = load_workbook(filename)
'''
def WriteAcceptanceRates():
    for wks in book.worksheets:
        for i in range(2, len(acceptance_rates)+2):
            wks.cell(row=i, column=3).value = acceptance_rates[i-2]
'''
for wks in book.worksheets:
    for i in range(2, len(acceptance_rates)+2):
        luck = random.random()*9+1
        wks.cell(row=i, column=6).value = 1/luck
book.save(filename)
book.close
