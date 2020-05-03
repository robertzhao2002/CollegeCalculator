import pandas as pd
from openpyxl import load_workbook
import random


filename = 'CollegeCalculator/Colleges.xlsx'
df = pd.read_excel(filename)
num_rows = df.shape
columns = df.columns
columnlist = list(columns)
luck_index = columnlist.index('Pure Luck (10%)')
book = load_workbook(filename)
for wks in book.worksheets:
    for i in range(2, num_rows[0]+2):
        luck = random.random()*4+1
        wks.cell(row=i, column=luck_index+1).value = 1/luck
book.save(filename)
book.close




exceldict = {

}
for i in columns:
    exceldict[i] = None

for c in columns:
    exceldict[c] = df[c].tolist()
